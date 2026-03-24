"""
Excel文件读取模块
负责解析Excel文件，支持多Sheet、多行表头、合并单元格
"""

import os
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.cleaner import clean_field_name
from utils.file_scanner import is_large_file


def optimize_dataframe_memory(df: pd.DataFrame) -> pd.DataFrame:
    """
    优化DataFrame内存使用
    - 数值类型downcast
    - 字符串转category（如果有大量重复值）
    """
    # 优化数值类型
    for col in df.select_dtypes(include=['int64']).columns:
        df[col] = pd.to_numeric(df[col], downcast='integer')
    
    for col in df.select_dtypes(include=['float64']).columns:
        df[col] = pd.to_numeric(df[col], downcast='float')
    
    # 字符串列：如果唯一值少于50%，转category
    for col in df.select_dtypes(include=['object']).columns:
        num_unique = df[col].nunique()
        num_total = len(df[col])
        if num_unique > 0 and num_unique / num_total < 0.5:
            df[col] = df[col].astype('category')
    
    return df


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, file_path: str):
        """
        初始化读取器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.filename = os.path.basename(file_path)
        self._sheets_info = None
        self._headers_cache = {}
    

    def get_memory_usage(self) -> Dict[str, Any]:
        """获取当前内存使用情况"""
        import psutil
        import sys
        
        process = psutil.Process()
        memory_info = process.memory_info()
        
        return {
            'rss_mb': memory_info.rss / 1024 / 1024,
            'vms_mb': memory_info.vms / 1024 / 1024,
            'percent': process.memory_percent()
        }

    def get_sheets(self) -> List[Dict[str, Any]]:
        """
        获取所有Sheet信息（大文件保护版本）
        
        Returns:
            Sheet信息列表 [{name, row_count, has_merged_cells}, ...]
        """
        if self._sheets_info is not None:
            return self._sheets_info
        
        self._sheets_info = []
        
        # 检查文件大小，超过10MB使用极简模式（原30MB太大）
        import os
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        
        # 大文件保护：即使不到10MB，如果行数很多也要保护
        if file_size_mb > 10:
            # 大文件只获取sheet名称，不读取任何数据
            try:
                # 使用 read_only 模式并关闭所有计算
                wb = load_workbook(self.file_path, read_only=True, data_only=False, 
                                   keep_vba=False, keep_links=False)
                for sheet_name in wb.sheetnames:
                    self._sheets_info.append({
                        'name': sheet_name,
                        'row_count': 0,  # 不读取，避免内存占用
                        'has_merged_cells': False
                    })
                wb.close()
                print(f"[ExcelReader] 大文件 {self.filename} ({file_size_mb:.1f}MB) 使用快速模式")
            except Exception as e:
                print(f"Error reading sheets: {e}")
                return []
            return self._sheets_info
        
        # 小文件正常读取
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_count = ws.max_row if ws.max_row else 0
                
                try:
                    has_merged = len(ws.merged_cells.ranges) > 0 if hasattr(ws, 'merged_cells') and ws.merged_cells else False
                except Exception:
                    has_merged = False
                
                self._sheets_info.append({
                    'name': sheet_name,
                    'row_count': row_count,
                    'has_merged_cells': has_merged
                })
            
            wb.close()
            
        except Exception as e:
            print(f"Error reading sheets: {e}")
            raise
        
        return self._sheets_info
    
    def get_headers(self, sheet_name: str, header_rows: int = 1) -> List[str]:
        """
        获取表头（支持多行表头和合并单元格）
        
        Args:
            sheet_name: Sheet名称
            header_rows: 表头行数
            
        Returns:
            表头列表
        """
        cache_key = f"{sheet_name}_{header_rows}"
        if cache_key in self._headers_cache:
            return self._headers_cache[cache_key]
        
        try:
            # 读取Excel，跳过空行
            # 检查文件大小
            import os
            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
            
            # 大文件保护：超过10MB直接用openpyxl读表头，不通过pandas
            if file_size_mb > 10:
                print(f"[ExcelReader] 大文件 {self.filename} ({file_size_mb:.1f}MB) 使用快速表头读取")
                wb = load_workbook(self.file_path, read_only=True, data_only=True, 
                                   keep_vba=False, keep_links=False)
                ws = wb[sheet_name]
                headers = []
                for cell in next(ws.iter_rows(min_row=1, max_row=header_rows, values_only=True)):
                    if cell:
                        headers.append(clean_field_name(str(cell)))
                wb.close()
                # 处理重复列名
                headers = self._handle_duplicate_headers(headers)
                self._headers_cache[cache_key] = headers
                return headers
            
            # 小文件使用pandas快速读取
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=list(range(header_rows)),  # 多行表头
                nrows=0,  # 只读表头
                engine='openpyxl'
            )
            
            # 处理多行表头 - 合并为单行
            headers = []
            
            # 获取所有列
            if isinstance(df.columns, pd.MultiIndex):
                # 多行表头 - 合并列名
                for col in df.columns:
                    # 取非空的最上层名称
                    col_name = col[0] if col[0] else ''
                    if not col_name:
                        for c in col:
                            if c:
                                col_name = c
                                break
                    headers.append(clean_field_name(col_name))
            else:
                # 单行表头
                headers = [clean_field_name(h) for h in df.columns.tolist()]
            
            # 处理重复列名
            headers = self._handle_duplicate_headers(headers)
            
            self._headers_cache[cache_key] = headers
            return headers
            
        except Exception as e:
            print(f"Error reading headers: {e}")
            raise
    
    def _handle_duplicate_headers(self, headers: List[str]) -> List[str]:
        """
        处理重复的列名
        
        Args:
            headers: 原始表头列表
            
        Returns:
            处理后的表头列表
        """
        seen = {}
        result = []
        
        for h in headers:
            if h in seen:
                seen[h] += 1
                result.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                result.append(h)
        
        return result
    
    def read_sheet(self, sheet_name: str, header_rows: int = 1, 
                   use_cache: bool = True, max_memory_mb: int = 300,
                   max_rows: int = None, usecols: list = None) -> pd.DataFrame:
        """
        读取Sheet数据（查询时调用，Dask处理大文件）
        
        Args:
            sheet_name: Sheet名称
            header_rows: 表头行数
            use_cache: 是否使用缓存
            max_memory_mb: 最大内存限制（MB）- Dask会自适应处理
            max_rows: 最大读取行数（用于大文件限制）
            usecols: 只读取指定的列名列表
            
        Returns:
            DataFrame
        """
        cache_key = f"{sheet_name}_{header_rows}_data"
        
        if use_cache and hasattr(self, '_data_cache') and cache_key in self._data_cache:
            return self._data_cache[cache_key]
        
        try:
            # 检查文件大小
            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
            
            # 大文件使用分块读取（阈值10MB）
            use_chunked = file_size_mb > 10
            
            print(f"[ExcelReader] 读取 {self.filename} - {sheet_name} ({file_size_mb:.1f}MB)" + 
                  (" [分块模式]" if use_chunked else ""))
            
            # 使用 openpyxl 的 read_only 模式读取，减少内存占用
            try:
                # 大文件必须使用分块读取
                if use_chunked:
                    return self._read_large_file(sheet_name, header_rows, max_rows)
                
                # 小文件正常读取
                read_kwargs = {
                    'sheet_name': sheet_name,
                    'header': list(range(header_rows)),
                    'engine': 'openpyxl',
                }
                if max_rows:
                    read_kwargs['nrows'] = max_rows
                if usecols:
                    read_kwargs['usecols'] = usecols
                    print(f"[ExcelReader] 只读取指定列: {usecols}")
                
                df = pd.read_excel(self.file_path, **read_kwargs)
                
                # 清洗列名
                df.columns = [clean_field_name(col) for col in df.columns]
                
                # 检查内存占用
                import sys
                memory_usage = df.memory_usage(deep=True).sum()
                memory_mb = memory_usage / (1024 * 1024)
                
                if memory_mb > max_memory_mb:
                    print(f"[ExcelReader] 警告：内存使用 {memory_mb:.1f}MB > {max_memory_mb}MB，尝试优化...")
                    
                    # 对于大内存占用，只保留必要的列
                    if all(col in df.columns for col in ['姓名', '部门']):  # 示例字段
                        essential_cols = [col for col in df.columns if 'ID' in col or '编号' in col]
                        df = df[essential_cols]
                
                # 限制缓存大小
                cache_size = 3  # 最多缓存3个表
                if not hasattr(self, '_data_cache'):
                    self._data_cache = {}
                
                if len(self._data_cache) >= cache_size:
                    # 移除最旧的缓存
                    oldest_key = next(iter(self._data_cache))
                    del self._data_cache[oldest_key]
                    print(f"[ExcelReader] 清除旧缓存: {oldest_key}")
                
                self._data_cache[cache_key] = df
                
                # 优化内存使用
                df = optimize_dataframe_memory(df)
                print(f"[ExcelReader] 内存优化后: {df.memory_usage(deep=True).sum() / 1024 / 1024:.1f}MB")
                
                return df
                
            except MemoryError:
                print(f"[ExcelReader] 内存不足，尝试分块读取...")
                # 分块读取作为备选方案
                return self._read_large_file(sheet_name, header_rows)
            
        except Exception as e:
            print(f"Error reading sheet data: {e}")
            raise
    
    def _read_large_file(self, sheet_name: str, header_rows: int, max_rows: int = None) -> pd.DataFrame:
        """
        读取大文件（使用Dask自动分块，更高效且不占内存）
        
        Args:
            sheet_name: Sheet名称
            header_rows: 表头行数
            max_rows: 最大读取行数
            
        Returns:
            DataFrame
        """
        print(f"[ExcelReader] 使用Dask读取大文件: {self.filename}" + 
              (f", 限制{max_rows}行" if max_rows else ""))
        
        try:
            import dask.dataframe as dd
            
            # 使用Dask读取Excel，自动分块处理
            ddf = dd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=list(range(header_rows)),
                engine='openpyxl'
            )
            
            # 限制行数（如果指定了max_rows）
            if max_rows:
                ddf = ddf.head(max_rows, npartitions=-1)
            
            # 计算并转换为pandas（在Dask内部优化内存）
            # Dask会自动分块计算，不会一次性加载所有数据
            df = ddf.compute() if hasattr(ddf, 'compute') else ddf
            
            # 清洗列名
            df.columns = [clean_field_name(col) for col in df.columns]
            
            print(f"[ExcelReader] Dask读取完成: {len(df)} 行")
            return df
            
        except ImportError:
            print("[ExcelReader] Dask未安装，使用openpyxl流式读取")
            # 回退到openpyxl流式读取
            return self._read_large_file_fallback(sheet_name, header_rows, max_rows)
        except Exception as e:
            print(f"[ExcelReader] Dask读取失败: {e}，使用备用方法")
            return self._read_large_file_fallback(sheet_name, header_rows, max_rows)
    
    def _read_large_file_fallback(self, sheet_name: str, header_rows: int, max_rows: int = None) -> pd.DataFrame:
        """
        备用大文件读取方法（openpyxl流式）
        """
        print(f"[ExcelReader] 使用openpyxl流式读取: {self.filename}" + 
              (f", 限制{max_rows}行" if max_rows else ""))
        
        wb = load_workbook(
            self.file_path, 
            read_only=True, 
            data_only=True,
            keep_vba=False, 
            keep_links=False
        )
        ws = wb[sheet_name]
        
        # 先读取表头
        header = []
        for i, row in enumerate(ws.iter_rows(max_row=header_rows, values_only=True)):
            if i == 0:
                header = [clean_field_name(str(c)) if c else f"col_{j}" for j, c in enumerate(row)]
                break
        
        # 流式读取数据
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows and i >= max_rows:
                break
            if i % 10000 == 0 and i > 0:
                print(f"[ExcelReader] 已读取 {i} 行...")
            rows.append(row)
        
        wb.close()
        
        if not rows:
            return pd.DataFrame(columns=header)
        
        df = pd.DataFrame(rows, columns=header[:len(rows[0])] if rows else header)
        print(f"[ExcelReader] 流式读取完成: {len(df)} 行")
        
        return df
    
    def get_field_info(self, sheet_name: str, max_rows: int = 5) -> List[Dict[str, str]]:
        """
        获取字段详细信息（只读取前N行，避免大文件崩溃）
        
        Args:
            sheet_name: Sheet名称
            max_rows: 最大读取行数，默认5行（减少内存占用）
            
        Returns:
            字段信息列表 [{name, type, nullable}, ...]
        """
        # 只读取前5行数据，避免大文件导致系统崩溃
        try:
            # 检查文件大小，超过10MB直接跳过字段类型检测（原50MB太大）
            import os
            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
            
            if file_size_mb > 10:
                # 大文件只返回基本字段名，跳过类型检测
                headers = self.get_headers(sheet_name, 1)
                return [{'name': h, 'type': 'unknown', 'nullable': True} for h in headers]
            
            # 使用 openpyxl 引擎
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                nrows=max_rows,
                engine='openpyxl'
            )
            
            # 清洗列名
            df.columns = [clean_field_name(col) for col in df.columns]
        except Exception as e:
            print(f"Error reading preview data: {e}")
            # 如果读取失败，返回空DataFrame
            df = pd.DataFrame()
        
        fields = []
        for col in df.columns:
            fields.append({
                'name': col,
                'type': str(df[col].dtype),
                'nullable': df[col].isna().any() if len(df) > 0 else True
            })
        
        return fields
    
    def preview_data(self, sheet_name: str, nrows: int = 10) -> pd.DataFrame:
        """
        预览数据（前N行）
        
        Args:
            sheet_name: Sheet名称
            nrows: 行数
            
        Returns:
            DataFrame
        """
        df = pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            nrows=nrows,
            engine='openpyxl'
        )
        
        # 清洗列名
        df.columns = [clean_field_name(col) for col in df.columns]
        
        return df
    
    def get_summary(self) -> Dict[str, Any]:
        """
        获取文件摘要信息
        
        Returns:
            摘要信息
        """
        size = os.path.getsize(self.file_path)
        
        return {
            'filename': self.filename,
            'path': self.file_path,
            'size': size,
            'size_mb': round(size / (1024 * 1024), 2),
            'sheet_count': len(self.get_sheets()),
            'sheets': self.get_sheets(),
            'is_large': is_large_file(self.file_path)
        }

def create_reader(file_path: str) -> ExcelReader:
    """
    工厂函数 - 创建ExcelReader实例
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        ExcelReader实例
    """
    return ExcelReader(file_path)